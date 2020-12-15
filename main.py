from openpyxl import load_workbook
from pip._vendor import requests
from time import sleep

#checking_columns = ['B', 'C', 'D', 'F', 'G', 'H']
#postal_code_column = ''
#long_code_column = 'J'
#lat_code_column = 'K'
#start_index = 1
#wb_name = 'Serviceability  List for Cable & EoC Sites-Febrary 2020.xlsx'

checking_columns = ['B']
postal_code_column = 'E'
long_code_column = 'F'
lat_code_column = 'G'
start_index = 1
wb_name = 'Toronto On-Net Buildings.xlsx'

def form_address(sheet, row, columns):
    address = ''
    for column in columns:
        address = "{0} {1}".format(address, sheet["{0}{1}".format(column, row + 1)].value)

    return address


def get_postal_code_and_coordinates(address, first_start):
    url = "{0}{1}".format('https://geocode.xyz/', address.replace(' ', '+'))
    params = {'json': 1}
    request = requests.get(url, params)
    json = request.json()
    if 'standard' not in json.keys():
        if not first_start:
            return {'postal_code': '',
                    'long': '',
                    'lat': ''}
        sleep(1)
        return get_postal_code_and_coordinates(address, False)
    code = str(json.get('standard', {}).get('postal', {}))
    if code == '{}':
        code = str(json.get('alt', {}).get('loc', {}).get('postal', {}))
    return {'postal_code': code,
            'long': json.get('longt', ''),
            'lat': json.get('latt', '')}

def parse():
    sheet = wb[wb.sheetnames[0]]
    if postal_code_column:
        sheet["{0}{1}".format(postal_code_column, 1)] = "Postal code"
    if long_code_column:
        sheet["{0}{1}".format(long_code_column, 1)] = "Long"
    if lat_code_column:
        sheet["{0}{1}".format(lat_code_column, 1)] = "Lat"

    for row in range(start_index, sheet.max_row):
        address = form_address(sheet, row, checking_columns)
        result = get_postal_code_and_coordinates(address, True)
        try:
            if postal_code_column:
                sheet["{0}{1}".format(postal_code_column, row + 1)] = result['postal_code']
            if long_code_column:
                sheet["{0}{1}".format(long_code_column, row + 1)] = result['long']
            if lat_code_column:
                sheet["{0}{1}".format(lat_code_column, row + 1)] = result['lat']
            print(result)
        except ValueError as error:
            wb.save(wb_name)
            print(error)
        sleep(0.6)
    wb.save(wb_name)


wb = load_workbook(wb_name)
try:
    parse()
except KeyboardInterrupt:
    wb.save(wb_name)
    print("exit")
